# -*- coding: utf-8 -*-
"""
rapor.py — Tarama sonuçlarından biçimli Excel raporu üretir (openpyxl).
Sayfalar: Sinyaller (hepsi), Kesin Al, Kesin Sat, Sektör Özeti.
"""
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import payload as pl

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="111418")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
AL_FILL = PatternFill("solid", fgColor="E3F7EC")
SAT_FILL = PatternFill("solid", fgColor="FCE4E4")
MAN_FILL = PatternFill("solid", fgColor="F0E8FC")
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("Kod", 10), ("Şirket", 22), ("Karar", 12), ("Fiyat (₺)", 11),
    ("AV Skoru", 9), ("Güven %", 9), ("Kazanma %", 10), ("Süre (gün)", 10),
    ("Giriş (₺)", 11), ("Hedef (₺)", 11), ("Stop (₺)", 11),
    ("Potansiyel %", 12), ("Risk %", 9), ("R/R", 7), ("RSI", 7),
    ("Akıllı Para", 11), ("Niyet", 18), ("Sektör", 20),
]

KARAR_ETIKET = {"simdi": "KESİN AL", "al": "AL", "izle": "İZLE", "kesinSat": "KESİN SAT"}


def _satir(u):
    karar = KARAR_ETIKET.get(u.get("v"), "")
    if "manip" in (u.get("lists") or []):
        karar = "MANİPÜLASYON"
    return [
        u["tk"], u["nm"], karar, u["px"],
        u["av"], u["guven"], pl.win_pct(u), pl.sure_gun(u),
        u["giris"], u["hedef"], u["stop"],
        u["kz"] if u["side"] == "AL" else -u["kz"], u["ky"], u["rr"], u["rsi"],
        u["sm"], u.get("niyet", ""), u.get("sektor", ""),
    ]


def _yaz_sayfa(ws, basliklar, rows):
    ws.sheet_view.showGridLines = False
    # başlık
    for c, (ad, w) in enumerate(COLS, 1):
        cell = ws.cell(1, c, ad)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    # satırlar
    for i, u in enumerate(rows, start=2):
        vals = _satir(u)
        man = "manip" in (u.get("lists") or [])
        fill = MAN_FILL if man else (AL_FILL if u["side"] == "AL" else SAT_FILL)
        for c, val in enumerate(vals, 1):
            cell = ws.cell(i, c, val)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.fill = fill
            if c in (4, 9, 10, 11):
                cell.number_format = '#,##0.00'
            elif c in (7, 12, 13):
                cell.number_format = '0.0'
            elif c == 14:
                cell.number_format = '0.0'
            cell.alignment = Alignment(horizontal="left" if c in (1, 2, 3, 17, 18) else "center")
        # karar hücresi vurgusu
        kc = ws.cell(i, 3)
        if u["side"] == "AL":
            kc.font = Font(name=FONT, size=10, bold=True, color="0A9E68")
        elif man:
            kc.font = Font(name=FONT, size=10, bold=True, color="7A3FCF")
        else:
            kc.font = Font(name=FONT, size=10, bold=True, color="C42B2B")


def _sektor_sayfa(ws, sectors):
    ws.sheet_view.showGridLines = False
    bsl = [("Sektör", 26), ("Ort. AV Skoru", 14), ("Akıllı Para", 13), ("Hisse Sayısı", 12), ("Isı", 10)]
    for c, (ad, w) in enumerate(bsl, 1):
        cell = ws.cell(1, c, ad)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    for i, s in enumerate(sectors, start=2):
        ort = s["ort"]
        isi = "🟢 Sıcak" if ort >= 65 else ("🟡 Ilık" if ort >= 50 else "🔴 Soğuk")
        renk = "E3F7EC" if ort >= 65 else ("FFF6E0" if ort >= 50 else "FCE4E4")
        vals = [s["sektor"], ort, s["sm"], s["n"], isi]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(i, c, val)
            cell.font = Font(name=FONT, size=10)
            cell.fill = PatternFill("solid", fgColor=renk)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left" if c in (1, 5) else "center")


def excel_rapor(sonuclar, out_path=None, xu100=None):
    """sonuclar: analiz_et + zenginleştirme listesi. Dönen: out_path."""
    if out_path is None:
        out_path = f"BIST_rapor_{datetime.date.today().isoformat()}.xlsx"
    uis = []
    for r in (sonuclar or []):
        try:
            uis.append(pl.to_ui(r))
        except Exception:
            continue
    uis.sort(key=lambda u: u.get("av", 0), reverse=True)

    wb = Workbook()
    ws = wb.active; ws.title = "Sinyaller"
    _yaz_sayfa(ws, COLS, uis)
    _yaz_sayfa(wb.create_sheet("Kesin Al"), COLS, [u for u in uis if u["side"] == "AL" and u["v"] in ("simdi", "al")])
    _yaz_sayfa(wb.create_sheet("Kesin Sat"), COLS, [u for u in uis if u["side"] == "SAT"])
    _sektor_sayfa(wb.create_sheet("Sektör Özeti"), pl.build_sectors(sonuclar))

    # üst bilgi notu (Sinyaller sayfasının altına)
    son = len(uis) + 3
    ws.cell(son, 1, f"BIST Para Avcısı · {datetime.datetime.now():%d.%m.%Y %H:%M} · Karar destek aracıdır, yatırım tavsiyesi değildir.").font = Font(name=FONT, size=9, italic=True, color="888888")

    wb.save(out_path)
    return out_path
